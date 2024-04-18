# flake8: noqa: E501
import csv
import os
import os.path as osp
import re
import shutil
from collections import defaultdict
from datetime import datetime
import json

import mmengine
import numpy as np
from mmengine import ConfigDict
import pandas as pd

try:
    from prettytable import from_csv
except ImportError:
    from_csv = None

from opencompass.utils import dataset_abbr_from_cfg, model_abbr_from_cfg
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

key = os.getenv('SENSENOVA_API_KEY')
headers = {
    "Authorization": f'Bearer {key}',
    "Content-Type": "application/json"
}

def download_file_by_fileid(file_id, fn_path):
    url = f"https://file.stage.sensenova.cn/v1/files/{file_id}/content"
    response = requests.get(url, headers=headers)

    if response.content.startswith(b'\xff\xd8'):
        with open(fn_path, "wb") as file:
            file.write(response.content)
    else:
        raise Exception('The content is not a JPEG image.')


def download_file_by_url(url, fn_path):
    response = requests.get(url)

    if response.status_code == 200:
        with open(fn_path, 'wb') as file:
            file.write(response.content)
    else:
        raise Exception('Failed to download image url: ', url)


class AssistantSTSummarizer:
    """Do the subjectivity analyze based on evaluation results.

    Args:
        config (ConfigDict): The configuration object of the evaluation task.
            It's expected to be filled out at runtime.
    """

    def __init__(self, config: ConfigDict) -> None:
        self.tasks = []
        self.cfg = config

    def summarize_excel(self, time_str: str = datetime.now().strftime('%Y%m%d_%H%M%S')):
        '''
        1. read predictions and output to csv
        '''

        dataset_cfgs = self.cfg['datasets']
        work_dir = self.cfg['work_dir']
        self.work_dir = work_dir

        self.time_str = time_str
        output_path = osp.join(self.work_dir, 'summary',
                               f'summary_{self.time_str}.txt')
        summary_dir = osp.join(osp.split(output_path)[0], f'{self.time_str}')
        mmengine.mkdir_or_exist(summary_dir)        

        # cp base excel to summary folder
        base_excel_path = './data/subjective/agent_base.xlsx'
        dst_excel_path = f'{summary_dir}/base.xlsx'
        os.system(f'cp {base_excel_path} {dst_excel_path}')

        self.out_excel_path = osp.join(summary_dir, 'summary.xlsx')

        results_folder = osp.join(work_dir, 'results')
        prediction_folder = osp.join(work_dir, 'predictions')

        for subdir in os.listdir(prediction_folder):
            model_abbr = subdir
            # mv picture_cache_dir to outputs
            self.picture_cache_dir = f'tmp/{model_abbr}'

            subdir_pred_path = os.path.join(prediction_folder, subdir)

            subdir_summary_path = os.path.join(summary_dir, subdir + '_files')
            mmengine.mkdir_or_exist(subdir_summary_path)        

            if os.path.isdir(subdir_pred_path):
                model = subdir
                for dataset in dataset_cfgs:
                    dataset_abbr = dataset_abbr_from_cfg(dataset)
                    pred_filepath = os.path.join(subdir_pred_path,
                                            dataset_abbr + '.json')
                    preds = mmengine.load(pred_filepath)

                    base_data = pd.read_excel(dst_excel_path)

                    # Load results
                    for k, v in preds.items():
                        index = int(k) + 1

                        prediction = preds[k]['prediction']

                        try:
                            tool_call = prediction['tool_call']
                            output_file_id = prediction['output_file_id']
                            output_file_url = prediction['output_file_url']

                            content = prediction['content']

                            row = base_data[base_data['序号'] == index]

                            base_data.loc[row.index, model_abbr] = content
                            if tool_call != 'None':
                                base_data.loc[row.index, 'tool_call'] = tool_call

                            if output_file_url != 'None':
                                src_image_path = os.path.join(self.picture_cache_dir,
                                                      f'{output_file_id}.jpg')
                                dst_image_path = os.path.join(subdir_summary_path, f'{index}-{output_file_id}.jpg')
                                try:
                                    # print(f'mv {src_image_path} {dst_image_path}')
                                    shutil.copy(src_image_path, dst_image_path)
                                except:
                                    print(f"missing file {index}-{output_file_id}")
                                base_data.loc[row.index, model_abbr] = output_file_url
                        except:
                            row = base_data[base_data['序号'] == index]
                            base_data.loc[row.index, model_abbr] = str(prediction)


                base_data.to_excel(self.out_excel_path)
                
                # self.add_image_to_excel()

    def add_image_to_excel(self):
        wb = load_workbook(self.out_excel_path)
        ws = wb['Sheet1']

        for fn in os.listdir(self.output_image_dir):
            index = int(fn.split('-')[0])
            img = Image(os.path.join(self.output_image_dir, fn))
            cell = 'content' + str(index)
            ws.add_image(img, cell)

        wb.save(self.out_excel_path)

    def summarize(self,
                time_str: str = datetime.now().strftime('%Y%m%d_%H%M%S')):
        self.summarize_excel(time_str)


class AssistantInternLM2Summarizer:
    """Do the subjectivity analyze based on evaluation results.

    Args:
        config (ConfigDict): The configuration object of the evaluation task.
            It's expected to be filled out at runtime.
    """

    def __init__(self, config: ConfigDict) -> None:
        self.tasks = []
        self.cfg = config

    def summarize_excel(self, time_str: str = datetime.now().strftime('%Y%m%d_%H%M%S')):
        '''
        1. read predictions and output to csv
        '''

        dataset_cfgs = self.cfg['datasets']
        work_dir = self.cfg['work_dir']
        self.work_dir = work_dir

        self.time_str = time_str
        output_path = osp.join(self.work_dir, 'summary',
                               f'summary_{self.time_str}.txt')
        summary_dir = osp.join(osp.split(output_path)[0], f'{self.time_str}')
        mmengine.mkdir_or_exist(summary_dir)        

        # cp base excel to summary folder
        base_excel_path = './data/subjective/agent_base.xlsx'
        dst_excel_path = f'{summary_dir}/base.xlsx'
        os.system(f'cp {base_excel_path} {dst_excel_path}')

        self.out_excel_path = osp.join(summary_dir, 'summary.xlsx')

        results_folder = osp.join(work_dir, 'results')
        prediction_folder = osp.join(work_dir, 'predictions')

        for subdir in os.listdir(prediction_folder):
            model_abbr = subdir
            # mv picture_cache_dir to outputs
            self.picture_cache_dir = f'tmp/{model_abbr}'

            subdir_pred_path = os.path.join(prediction_folder, subdir)

            subdir_summary_path = os.path.join(summary_dir, subdir + '_files')
            mmengine.mkdir_or_exist(subdir_summary_path)        

            if os.path.isdir(subdir_pred_path):
                model = subdir
                for dataset in dataset_cfgs:
                    dataset_abbr = dataset_abbr_from_cfg(dataset)
                    pred_filepath = os.path.join(subdir_pred_path,
                                            dataset_abbr + '.json')
                    preds = mmengine.load(pred_filepath)

                    base_data = pd.read_excel(dst_excel_path)

                    # Load results
                    for k, v in preds.items():
                        index = int(k) + 1

                        prediction = preds[k]['prediction']

                        try:
                            tool_call = prediction[0]['name']
                            # arguements = prediction['arguments']

                            row = base_data[base_data['序号'] == index]

                            base_data.loc[row.index, model_abbr] = str(prediction)
                            base_data.loc[row.index, 'tool_call'] = tool_call

                        except:
                            row = base_data[base_data['序号'] == index]
                            base_data.loc[row.index, model_abbr] = str(prediction)


                base_data.to_excel(self.out_excel_path)
                
                # self.add_image_to_excel()

    def add_image_to_excel(self):
        wb = load_workbook(self.out_excel_path)
        ws = wb['Sheet1']

        for fn in os.listdir(self.output_image_dir):
            index = int(fn.split('-')[0])
            img = Image(os.path.join(self.output_image_dir, fn))
            cell = 'content' + str(index)
            ws.add_image(img, cell)

        wb.save(self.out_excel_path)

    def summarize(self,
                time_str: str = datetime.now().strftime('%Y%m%d_%H%M%S')):
        self.summarize_excel(time_str)